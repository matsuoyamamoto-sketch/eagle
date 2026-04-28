"""EDC仕様書 (Excel) 生成の検証。"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.generators.spec_excel import write_spec_excel
from src.parser.edc_parser import load_study

SAMPLE_JSON = Path(
    r"c:/Users/matsu/Box/Datacenter/ISR/Ptosh/検証/JSON/Bev-FOLFOX-SBC_250929_1501.json"
)


@pytest.fixture(scope="module")
def spec_path(tmp_path_factory) -> Path:
    if not SAMPLE_JSON.exists():
        pytest.skip("sample JSON not found")
    study = load_study(SAMPLE_JSON)
    out = tmp_path_factory.mktemp("spec") / "edc_spec.xlsx"
    return write_spec_excel(study, out)


def test_workbook_structure(spec_path: Path):
    wb = load_workbook(spec_path, read_only=True)
    names = wb.sheetnames
    assert names[0] == "表紙"
    assert names[1] == "改訂履歴"
    assert names[2] == "00_索引"
    assert names[3] == "00_シートマトリクス"
    assert names[4] == "00_コードリスト一覧"
    # 5固定シート + 162フォーム
    assert len(names) == 5 + 162


def test_cover_has_study_name(spec_path: Path):
    wb = load_workbook(spec_path, read_only=True)
    ws = wb["表紙"]
    rows = list(ws.iter_rows(values_only=True))
    flat = [c for r in rows for c in r if c]
    assert "Bev-FOLFOX-SBC" in flat


def test_index_meta(spec_path: Path):
    wb = load_workbook(spec_path, read_only=True)
    ws = wb["00_索引"]
    rows = list(ws.iter_rows(values_only=True))
    metas = {r[0]: r[1] for r in rows if r and r[0] and r[1] is not None}
    assert metas.get("試験ID") == "Bev-FOLFOX-SBC"
    assert metas.get("フォーム数") == 162
    assert metas.get("総入力項目数") == 32716


def test_codelist_only_used(spec_path: Path):
    """使用中のコードリストのみが書き出されていること。"""
    wb = load_workbook(spec_path, read_only=True)
    ws = wb["00_コードリスト一覧"]
    text = " ".join(
        str(c) for r in ws.iter_rows(values_only=True) for c in r if c
    )
    # 索引メタの「使用中」件数が 85(全体) より少ないか同じ
    idx = wb["00_索引"]
    used_count = None
    for r in idx.iter_rows(values_only=True):
        if r and r[0] == "コードリスト数 (使用中)":
            used_count = r[1]
            break
    assert used_count is not None
    assert used_count <= 85
    # 必ず参照されるはずの "Sex" は含まれる
    assert "Sex" in text


def test_form_sheet_layout(spec_path: Path):
    """A1=シート名, A2=索引リンク, A4=表ヘッダー, B列=Field, C列=SDTM。"""
    wb = load_workbook(spec_path, read_only=True)
    sheet_name = wb.sheetnames[5]  # 5 固定 + 最初のフォーム
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # A1
    assert "症例登録票" in str(rows[0][0])
    # A2
    assert "索引" in str(rows[1][0])
    # A4 ヘッダー: A=Seq, B=Field, C=SDTM, D=Label
    header = rows[3]
    assert header[0] == "Seq"
    assert header[1] == "Field"
    assert header[2] == "SDTM"
    assert header[3] == "Label"
    # データ: field2 (同意取得日) の SDTM=DM.RFICDTC
    found = False
    for r in rows[4:]:
        if r[1] == "field2":
            assert r[2] == "DM.RFICDTC"
            found = True
            break
    assert found, "field2 (同意取得日) が見つからない"


def test_matrix_sheet(spec_path: Path):
    """シートマトリクスが3群+未分類を持ち、合計が162になること。"""
    wb = load_workbook(spec_path, read_only=True)
    ws = wb["00_シートマトリクス"]
    rows = list(ws.iter_rows(values_only=True))
    # ヘッダーは row index 3 (A4)
    header = rows[3]
    assert header[0] == "No."
    assert header[1] == "Sheet Name"
    assert "未分類" in header
    # 集計行: 群は重複所属あり。未分類が 94, 各群の合計≧162。
    summary = rows[3 + 162 + 1]
    nums = [v for v in summary if isinstance(v, int)]
    assert sum(nums) >= 162
    # 未分類は最終列
    assert nums[-1] == 94
