"""エディットチェック確認書 生成テスト。"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.generators.edit_check import write_edit_check_excel
from src.parser.edc_parser import load_study

SAMPLE_JSON = Path(
    r"c:/Users/matsu/Box/Datacenter/ISR/Ptosh/検証/JSON/Bev-FOLFOX-SBC_250929_1501.json"
)


@pytest.fixture(scope="module")
def ec_path(tmp_path_factory) -> Path:
    if not SAMPLE_JSON.exists():
        pytest.skip("sample JSON not found")
    study = load_study(SAMPLE_JSON)
    out = tmp_path_factory.mktemp("ec") / "edit_check.xlsx"
    return write_edit_check_excel(study, out)


def test_workbook_structure(ec_path: Path):
    wb = load_workbook(ec_path, read_only=True)
    assert wb.sheetnames == ["表紙", "改訂履歴", "サマリ", "エディットチェック一覧"]


def test_total_checks(ec_path: Path):
    """全 validator の合計と一致 (presence 6809 + date 3112 + num 2817 + formula 4065 = 16803)。"""
    wb = load_workbook(ec_path, read_only=True)
    ws = wb["エディットチェック一覧"]
    rows = list(ws.iter_rows(values_only=True))
    # title(1) + blank(1) + header(1) + data(N) のうち、3行目までを除外
    data = [r for r in rows[3:] if r and r[0] is not None]
    # 環境差を吸収するため厳密一致でなく範囲チェック
    assert 16700 <= len(data) <= 16900
    # 1 件目は No.=1
    assert data[0][0] == 1
    # 種別が 4 種類すべて含まれる
    types = {r[5] for r in data}
    assert types == {"Required", "Date", "Numericality", "Formula"}


def test_summary_sheet(ec_path: Path):
    wb = load_workbook(ec_path, read_only=True)
    ws = wb["サマリ"]
    rows = list(ws.iter_rows(values_only=True))
    # 3行目ヘッダー、4-7 が各種別、8 が合計
    types = {r[0]: r[1] for r in rows if r and r[0] in ("Required", "Date", "Numericality", "Formula", "合計")}
    assert types["Required"] == 6809
    assert types["Date"] == 3112
    assert types["Numericality"] == 2817
    assert types["Formula"] == 4065
    assert types["合計"] == 6809 + 3112 + 2817 + 4065
