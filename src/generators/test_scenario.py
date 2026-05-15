"""テストシナリオ (Excel) 生成 — Cohere 利用。"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins, PrintOptions

from ..ai.cohere_client import CohereJSONClient
from ..ai.prompts import test_scenario as P
from ..parser.models import Sheet, Study

BASE_FONT = "Meiryo UI"
F_BASE = Font(name=BASE_FONT, size=9)
F_HEADER = Font(name=BASE_FONT, size=10, bold=True, color="FFFFFF")
F_TITLE = Font(name=BASE_FONT, size=16, bold=True)
FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_NORMAL = PatternFill("solid", fgColor="E2EFDA")
FILL_ABNORMAL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS = ["No.", "Sheet", "Field", "Label", "Kind", "Input", "Expected", "Rationale", "チェック", "確認者", "確認日"]
WIDTHS = [6, 24, 12, 24, 10, 28, 36, 40, 8, 16, 12]


def _estimate_row_height(values: list, widths: list[int]) -> float:
    """各セルの折り返し行数を推定し、最大値から行高 (pt) を算出。"""
    max_lines = 1
    for v, w in zip(values, widths):
        text = str(v) if v is not None else ""
        if not text:
            continue
        # 改行を考慮
        for line in text.replace("\r", "").split("\n"):
            # 全角は概ね 2 char 幅相当として、列幅 (chars) を基準に行数推定
            est_chars = max(1, w - 1)
            # 日本語は1文字≒2幅。安全側で 1.5 倍に。
            wrap_lines = max(1, -(-int(len(line) * 1.5) // est_chars))
            max_lines = max(max_lines, wrap_lines)
    return max(15.0, min(150.0, max_lines * 14.0))


def _row_fill(kind: str) -> PatternFill | None:
    if kind == "normal":
        return FILL_NORMAL
    if kind == "abnormal":
        return FILL_ABNORMAL
    return None


def generate_scenarios(
    study: Study,
    selected_sheet_names: list[str],
    client: CohereJSONClient | None = None,
    on_progress: Callable[[int, int, str], None] | None = None,
    on_event: Callable[[dict], None] | None = None,
) -> list[dict]:
    """指定フォームについて AI でシナリオ生成。"""
    import time as _time

    client = client or CohereJSONClient()
    if on_event:
        client.event_hook = on_event
    target = [s for s in study.sheets if s.name in selected_sheet_names]
    out: list[dict] = []
    total = len(target)
    for i, sheet in enumerate(target, start=1):
        if on_progress:
            on_progress(i, total, sheet.name)
        if on_event:
            on_event({"phase": "sheet_start", "i": i, "total": total,
                      "name": sheet.name, "kind": "scenario"})
        t0 = _time.monotonic()
        try:
            data = client.chat_json(P.SYSTEM, P.build_user_prompt(sheet, study), P.SCHEMA)
            n_added = 0
            for sc in data.get("scenarios", []):
                out.append({"sheet": sheet.name, **sc})
                n_added += 1
            if on_event:
                on_event({"phase": "sheet_end", "i": i, "total": total,
                          "name": sheet.name, "kind": "scenario",
                          "elapsed": _time.monotonic() - t0, "items": n_added})
        except Exception as e:
            out.append(
                {
                    "sheet": sheet.name,
                    "field": "(error)",
                    "label": "",
                    "kind": "abnormal",
                    "input_value": "",
                    "expected_result": "",
                    "rationale": f"生成エラー: {e}",
                }
            )
            if on_event:
                on_event({"phase": "sheet_error", "i": i, "total": total,
                          "name": sheet.name, "kind": "scenario",
                          "elapsed": _time.monotonic() - t0, "error": str(e)[:200]})
    return out


def build_test_scenario_workbook(
    study: Study, scenarios: list[dict]
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # 表紙 (簡略)
    cv = wb.create_sheet("表紙")
    cv.sheet_view.showGridLines = False
    cv.column_dimensions["A"].width = 2
    cv.column_dimensions["B"].width = 22
    cv.column_dimensions["C"].width = 60
    cv.merge_cells("B3:C3")
    t = cv["B3"]
    t.value = "テストシナリオ"
    t.font = Font(name=BASE_FONT, size=24, bold=True, color="1F3864")
    t.alignment = CENTER
    cv.row_dimensions[3].height = 50
    cv.row_dimensions[5].height = 60
    cv.merge_cells("B5:C5")
    pn = cv["B5"]
    pn.value = study.proper_name
    pn.font = Font(name=BASE_FONT, size=12, color="404040")
    pn.alignment = CENTER
    metas = [("試験 ID", study.name), ("シナリオ件数", f"{len(scenarios):,}"), ("発行日", date.today().isoformat())]
    for i, (k, v) in enumerate(metas):
        r = 9 + i
        cv.row_dimensions[r].height = 22
        kc = cv.cell(row=r, column=2, value=k)
        kc.font = Font(name=BASE_FONT, size=10, bold=True)
        kc.fill = PatternFill("solid", fgColor="D9E1F2")
        kc.border = BORDER_ALL
        kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc = cv.cell(row=r, column=3, value=v)
        vc.font = F_BASE
        vc.border = BORDER_ALL
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # 一覧シート
    ws = wb.create_sheet("テストシナリオ一覧")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "テストシナリオ一覧"
    ws["A1"].font = F_TITLE

    # 操作ガイド (タイトル下)
    ws["A2"] = (
        "📋 確認操作:  チェック列のドロップダウンから ✓ を選択 → "
        "確認者セルに名前を入力 → 確認日セルで Ctrl + ; (セミコロン) を押すと今日の日付が入ります。"
    )
    ws["A2"].font = Font(name=BASE_FONT, size=9, italic=True, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    header_row = 3
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]

    for idx, sc in enumerate(scenarios, start=1):
        r = header_row + idx
        values = [
            idx,
            sc.get("sheet", ""),
            sc.get("field", ""),
            sc.get("label", ""),
            sc.get("kind", ""),
            sc.get("input_value", ""),
            sc.get("expected_result", ""),
            sc.get("rationale", ""),
            "",  # チェック
            "",  # 確認者
            "",  # 確認日
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = F_BASE
            c.border = BORDER_ALL
            c.alignment = WRAP
        fill = _row_fill(sc.get("kind", ""))
        if fill is not None:
            ws.cell(row=r, column=5).fill = fill
        ws.row_dimensions[r].height = _estimate_row_height(values, WIDTHS)

    # チェック列にドロップダウン (✓)。VBA が変更を検知して確認者/確認日を自動入力。
    if scenarios:
        chk_col_letter = get_column_letter(HEADERS.index("チェック") + 1)
        first_data_row = header_row + 1
        last_data_row = header_row + len(scenarios)
        dv = DataValidation(type="list", formula1='"✓"', allow_blank=True,
                            showDropDown=False)
        dv.add(f"{chk_col_letter}{first_data_row}:{chk_col_letter}{last_data_row}")
        ws.add_data_validation(dv)
        # チェック列を中央揃え
        for r in range(first_data_row, last_data_row + 1):
            ws.cell(row=r, column=HEADERS.index("チェック") + 1).alignment = CENTER

    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(scenarios)}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    _build_completion_report(wb, study, scenarios)

    return wb


def _build_completion_report(wb: Workbook, study: Study, scenarios: list[dict]) -> None:
    """確認完了報告書シート (1ページ収まる印刷レイアウト)。"""
    rs = wb.create_sheet("確認完了報告書")
    rs.sheet_view.showGridLines = False

    # 印刷設定: A4 縦・1ページに収める
    rs.page_setup.paperSize = rs.PAPERSIZE_A4
    rs.page_setup.orientation = rs.ORIENTATION_PORTRAIT
    rs.page_setup.fitToWidth = 1
    rs.page_setup.fitToHeight = 1
    rs.sheet_properties.pageSetUpPr.fitToPage = True
    rs.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.5,
                                  header=0.3, footer=0.3)
    rs.print_options = PrintOptions(horizontalCentered=True)

    # 列幅 (A=2 余白、B〜F でレイアウト)。B 列は 210px (≒ 29.29 char)
    widths = {"A": 2, "B": 29.29, "C": 22, "D": 14, "E": 22, "F": 14}
    for col, w in widths.items():
        rs.column_dimensions[col].width = w

    F_TITLE_BIG = Font(name=BASE_FONT, size=18, bold=True, color="1F3864")
    F_LABEL = Font(name=BASE_FONT, size=10, bold=True, color="1F3864")
    F_VAL = Font(name=BASE_FONT, size=10)
    F_KPI_LABEL = Font(name=BASE_FONT, size=9, color="595959")
    F_KPI_VAL = Font(name=BASE_FONT, size=20, bold=True, color="1F3864")
    F_NOTE = Font(name=BASE_FONT, size=9, italic=True, color="595959")
    FILL_LBL = PatternFill("solid", fgColor="D9E1F2")
    FILL_KPI = PatternFill("solid", fgColor="F2F2F2")
    FILL_SECTION = PatternFill("solid", fgColor="1F3864")
    F_SECTION = Font(name=BASE_FONT, size=10, bold=True, color="FFFFFF")
    LEFT_C = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    CENTER_W = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def apply_border(r1: int, c1: int, r2: int, c2: int) -> None:
        """指定範囲内のすべてのセルに罫線を適用 (結合セル対策)。"""
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                rs.cell(row=r, column=c).border = BORDER_ALL

    def apply_fill(r1: int, c1: int, r2: int, c2: int, fill: PatternFill) -> None:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                rs.cell(row=r, column=c).fill = fill

    def estimate_lines(text: str, col_chars: float) -> int:
        """日本語混じりテキストの折り返し行数を推定。"""
        if not text:
            return 1
        max_lines = 0
        for line in str(text).replace("\r", "").split("\n"):
            wrap = max(1, -(-int(len(line) * 1.7) // max(1, int(col_chars))))
            max_lines += wrap
        return max(1, max_lines)

    # タイトル
    rs.merge_cells("B2:F2")
    rs["B2"] = "テストシナリオ 確認完了報告書"
    rs["B2"].font = F_TITLE_BIG
    rs["B2"].alignment = Alignment(horizontal="center", vertical="center")
    rs.row_dimensions[2].height = 32

    rs.merge_cells("B3:F3")
    rs["B3"] = study.proper_name or ""
    rs["B3"].font = Font(name=BASE_FONT, size=10, color="595959")
    rs["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # 試験名(proper_name)の長さに応じて行高を調整 (B3:F3 結合幅 ≒ 86 chars)
    proper_lines = estimate_lines(study.proper_name or "", 86)
    rs.row_dimensions[3].height = max(18, min(60, proper_lines * 14))

    # 試験情報
    def section(row: int, label: str) -> None:
        rs.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c = rs.cell(row=row, column=2, value=label)
        c.font = F_SECTION
        c.fill = FILL_SECTION
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rs.row_dimensions[row].height = 18

    def kv_row(row: int, k1: str, v1: str, k2: str | None = None, v2: str = "") -> None:
        # B: ラベル1
        rs.cell(row=row, column=2, value=k1).font = F_LABEL
        rs.cell(row=row, column=2).fill = FILL_LBL
        rs.cell(row=row, column=2).alignment = LEFT_C
        # C〜: 値1
        rs.cell(row=row, column=3, value=v1).font = F_VAL
        rs.cell(row=row, column=3).alignment = LEFT_C
        if k2 is not None:
            # D: ラベル2
            rs.cell(row=row, column=4, value=k2).font = F_LABEL
            rs.cell(row=row, column=4).fill = FILL_LBL
            rs.cell(row=row, column=4).alignment = LEFT_C
            # E:F 値2
            rs.cell(row=row, column=5, value=v2).font = F_VAL
            rs.cell(row=row, column=5).alignment = LEFT_C
            rs.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            value_chars = 22
        else:
            # C:F 値1
            rs.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            value_chars = 70  # C列22 + D14 + E22 + F14
        # 結合範囲も含めて全セルに罫線
        apply_border(row, 2, row, 6)
        # 値の長さに応じた行高
        max_text = max(str(v1 or ""), (str(v2) if k2 else ""), key=len)
        lines = estimate_lines(max_text, value_chars)
        rs.row_dimensions[row].height = max(20, min(80, lines * 16))

    section(5, "  ■ 試験情報")
    kv_row(6, "試験 ID", study.name, "発行日", date.today().isoformat())
    kv_row(7, "試験名",  study.proper_name or "-")
    kv_row(8, "対象シナリオ範囲", "全件 (一覧シート参照)")

    # KPI (件数サマリ)
    section(10, "  ■ 確認サマリ")
    total = len(scenarios)
    chk_col_letter = get_column_letter(HEADERS.index("チェック") + 1)
    src_first = 4  # header_row(3) + 1
    src_last = 3 + max(total, 1)
    src_range = f"テストシナリオ一覧!{chk_col_letter}{src_first}:{chk_col_letter}{src_last}"

    def kpi(col_start: str, col_end: str, row_label: int, row_value: int,
            label: str, value) -> None:
        from openpyxl.utils import column_index_from_string
        c1 = column_index_from_string(col_start)
        c2 = column_index_from_string(col_end)
        if c2 > c1:
            rs.merge_cells(start_row=row_label, start_column=c1, end_row=row_label, end_column=c2)
            rs.merge_cells(start_row=row_value, start_column=c1, end_row=row_value, end_column=c2)
        cl = rs[f"{col_start}{row_label}"]
        cl.value = label
        cl.font = F_KPI_LABEL
        cl.alignment = CENTER_W
        cv = rs[f"{col_start}{row_value}"]
        cv.value = value
        cv.font = F_KPI_VAL
        cv.alignment = CENTER_W
        # 範囲全体に塗りと罫線
        apply_fill(row_label, c1, row_label, c2, FILL_KPI)
        apply_fill(row_value, c1, row_value, c2, FILL_KPI)
        apply_border(row_label, c1, row_label, c2)
        apply_border(row_value, c1, row_value, c2)
        rs.row_dimensions[row_label].height = 16
        rs.row_dimensions[row_value].height = 30

    kpi("B", "C", 11, 12, "総シナリオ件数", total)
    kpi("D", "D", 11, 12, "確認済み", f'=COUNTIF({src_range},"✓")')
    kpi("E", "E", 11, 12, "未確認", f'=COUNTBLANK({src_range})')
    kpi("F", "F", 11, 12, "達成率",
        f'=IF({total}=0,"-",TEXT(COUNTIF({src_range},"✓")/{total},"0.0%"))')

    # 判定
    section(14, "  ■ 全件通過判定")
    rs.cell(row=15, column=2, value="判定結果").font = F_LABEL
    rs.cell(row=15, column=2).fill = FILL_LBL
    rs.cell(row=15, column=2).alignment = LEFT_C
    rs.merge_cells("C15:F15")
    judge_cell = rs["C15"]
    judge_cell.value = (
        f'=IF(COUNTIF({src_range},"✓")={total},"✅ 全件通過","⚠ 未通過あり ("'
        f'&({total}-COUNTIF({src_range},"✓"))&"件)")'
    )
    judge_cell.font = Font(name=BASE_FONT, size=12, bold=True)
    judge_cell.alignment = LEFT_C
    apply_border(15, 2, 15, 6)
    rs.row_dimensions[15].height = 24

    # ドロップダウン (総合判定の手動上書き用)
    rs.cell(row=16, column=2, value="最終承認判定").font = F_LABEL
    rs.cell(row=16, column=2).fill = FILL_LBL
    rs.cell(row=16, column=2).alignment = LEFT_C
    rs.merge_cells("C16:F16")
    rs["C16"].alignment = LEFT_C
    rs["C16"].font = F_VAL
    apply_border(16, 2, 16, 6)
    dv_judge = DataValidation(type="list",
                              formula1='"承認 (合格),条件付き承認,差戻し"',
                              allow_blank=True)
    dv_judge.add("C16")
    rs.add_data_validation(dv_judge)
    rs.row_dimensions[16].height = 22

    # 未通過項目の説明欄
    section(18, "  ■ 未チェック項目の理由 / 補足")
    rs.merge_cells("B19:F19")
    note_label = rs["B19"]
    note_label.value = (
        "（未チェックのまま許容した項目について、対象シナリオNo.と理由を記載してください。例: No.12 該当画面なし、運用上影響なし 等）"
    )
    note_label.font = F_NOTE
    note_label.alignment = LEFT_C
    rs.row_dimensions[19].height = 16

    rs.merge_cells("B20:F25")
    note_box = rs["B20"]
    note_box.value = ""
    note_box.alignment = Alignment(horizontal="left", vertical="top",
                                   wrap_text=True, indent=1)
    apply_border(20, 2, 25, 6)
    for r in range(20, 26):
        rs.row_dimensions[r].height = 22

    # 署名欄 (確認実施者は複数、承認者は1名)
    section(27, "  ■ 署名")

    # サブヘッダ: 区分 | 氏名 | 日付 | 署名/印
    sub_header_row = 28
    sub_headers = [("B", "区分"), ("C", "氏名"), ("D", "日付"), ("E", "署名 / 印")]
    for col, lbl in sub_headers:
        c = rs[f"{col}{sub_header_row}"]
        c.value = lbl
        c.font = F_LABEL
        c.fill = FILL_LBL
        c.alignment = CENTER_W
    # E列とF列を結合 (署名/印を広めに)
    rs.merge_cells(start_row=sub_header_row, start_column=5,
                   end_row=sub_header_row, end_column=6)
    apply_fill(sub_header_row, 5, sub_header_row, 6, FILL_LBL)
    apply_border(sub_header_row, 2, sub_header_row, 6)
    rs.row_dimensions[sub_header_row].height = 20

    # 確認実施者 (3名分)、承認者 (1名)
    sig_rows: list[tuple[str, str]] = [
        ("確認実施者 1", "verifier"),
        ("確認実施者 2", "verifier"),
        ("確認実施者 3", "verifier"),
        ("承認者",       "approver"),
    ]
    row = sub_header_row + 1
    for label, kind in sig_rows:
        # 区分ラベル
        lc = rs.cell(row=row, column=2, value=label)
        lc.font = F_LABEL if kind == "approver" else F_VAL
        row_fill = FILL_LBL if kind == "approver" else PatternFill("solid", fgColor="F8F9FA")
        lc.fill = row_fill
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # 氏名
        rs.cell(row=row, column=3, value="").font = F_VAL
        rs.cell(row=row, column=3).alignment = LEFT_C
        # 日付
        rs.cell(row=row, column=4, value="").font = F_VAL
        rs.cell(row=row, column=4).alignment = LEFT_C
        # 署名/印 (E:F 結合)
        rs.cell(row=row, column=5, value="").font = F_VAL
        rs.cell(row=row, column=5).alignment = LEFT_C
        rs.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        # 罫線を行全体に適用
        apply_border(row, 2, row, 6)
        rs.row_dimensions[row].height = 26
        row += 1

    last_sig_row = row - 1

    # 印刷範囲
    rs.print_area = f"B2:F{last_sig_row}"


def write_test_scenario_excel(
    study: Study,
    selected_sheet_names: list[str],
    output_path: str | Path,
    client: CohereJSONClient | None = None,
    on_progress: Callable[[int, int, str], None] | None = None,
) -> Path:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    scenarios = generate_scenarios(study, selected_sheet_names, client, on_progress)
    wb = build_test_scenario_workbook(study, scenarios)
    wb.save(out)
    return out
