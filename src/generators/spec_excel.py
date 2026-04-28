"""EDC仕様書 (Excel) 生成モジュール。

シート構成:
    1. 表紙
    2. 改訂履歴
    3. 00_索引
    4. 00_シートマトリクス (フォーム × 割付群)
    5. 00_コードリスト一覧 (実際にフォームから参照されているもののみ)
    6. 各フォーム (162)

各フォームシート:
    A1 = シート名
    A2 = ← 索引へ戻る リンク
    A4 から項目テーブル開始
"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from ..parser.models import FieldItem, Sheet, Study, Validators

# ---------------- スタイル ----------------
BASE_FONT_NAME = "Meiryo UI"

FONT_BASE = Font(name=BASE_FONT_NAME, size=9)
FONT_BOLD = Font(name=BASE_FONT_NAME, size=10, bold=True)
FONT_HEADER = Font(name=BASE_FONT_NAME, size=10, bold=True, color="FFFFFF")
FONT_TITLE = Font(name=BASE_FONT_NAME, size=16, bold=True)
FONT_SHEET_TITLE = Font(name=BASE_FONT_NAME, size=14, bold=True)
FONT_LINK = Font(name=BASE_FONT_NAME, size=10, color="0563C1", underline="single")
FONT_META_KEY = Font(name=BASE_FONT_NAME, size=10, bold=True)

FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_META = PatternFill("solid", fgColor="D9E1F2")
FILL_CODELIST_HEADER = PatternFill("solid", fgColor="70AD47")

THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_TOP_WRAP = Alignment(vertical="top", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")

# ---------------- 列定義 ----------------
FORM_HEADERS = [
    "Seq",
    "Field",
    "SDTM",
    "Label",
    "Type",
    "Field Type",
    "Required",
    "Visible",
    "Default",
    "Code List",
    "Min",
    "Max",
    "Date Min",
    "Date Max",
    "Formula (条件)",
    "Formula (メッセージ)",
    "Reference",
    "Description / Note",
]
FORM_COL_WIDTHS = [
    6, 12, 14, 28, 10, 12, 8, 8, 14, 22, 8, 8, 14, 14, 36, 28, 20, 40,
]

INDEX_HEADERS = ["No.", "Sheet Name", "Alias", "Category", "Items", "Closed", "Link"]


# ---------------- ユーティリティ ----------------
def _safe_sheet_name(raw: str, used: set[str], idx: int) -> str:
    name = _INVALID_SHEET_CHARS.sub("_", raw or f"Sheet{idx}").strip() or f"Sheet{idx}"
    prefix = f"{idx:03d}_"
    name = (prefix + name)[:31]
    base = name
    n = 1
    while name in used:
        suffix = f"_{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def _required(v: Validators) -> str:
    return "○" if v.presence is not None else ""


def _num_range(v: Validators) -> tuple[str, str]:
    if not v.numericality:
        return "", ""
    n = v.numericality
    lo = n.validate_numericality_greater_than_or_equal_to or n.validate_numericality_greater_than or ""
    hi = n.validate_numericality_less_than_or_equal_to or n.validate_numericality_less_than or ""
    if n.validate_numericality_equal_to:
        return n.validate_numericality_equal_to, n.validate_numericality_equal_to
    return lo, hi


def _date_range(v: Validators) -> tuple[str, str]:
    if not v.date:
        return "", ""
    d = v.date
    lo = d.validate_date_after_or_equal_to or d.validate_date_after or ""
    hi = d.validate_date_before_or_equal_to or d.validate_date_before or ""
    return lo, hi


def _formula(v: Validators) -> tuple[str, str]:
    if not v.formula:
        return "", ""
    return v.formula.validate_formula_if or "", v.formula.validate_formula_message or ""


def _description(fi: FieldItem) -> str:
    parts = []
    if fi.description:
        parts.append(fi.description)
    if fi.content:
        parts.append(fi.content)
    return "\n".join(parts)


def _build_sdtm_map(sheet: Sheet) -> dict[str, str]:
    """field名 → 'PREFIX.VAR'（複数あればカンマ結合）。"""
    out: dict[str, list[str]] = {}
    for cfg in sheet.cdisc_sheet_configs:
        prefix = cfg.prefix or ""
        for fname, var in (cfg.table or {}).items():
            if not var:
                continue
            label = f"{prefix}.{var}" if prefix else var
            out.setdefault(fname, []).append(label)
    return {k: ", ".join(v) for k, v in out.items()}


def _used_codelist_names(study: Study) -> set[str]:
    used: set[str] = set()
    for s in study.sheets:
        for fi in s.field_items:
            if fi.option_name:
                used.add(fi.option_name)
    return used


# ---------------- 表紙 ----------------
def _build_cover_sheet(wb: Workbook, study: Study) -> None:
    """A4縦・1ページに収まる印刷向け表紙。"""
    from openpyxl.worksheet.page import PageMargins

    ws = wb.create_sheet("表紙", index=0)
    ws.sheet_view.showGridLines = False

    # 印刷設定: A4縦, 1ページに収める, 中央寄せ, 余白
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.8, bottom=0.8)

    # 列幅: B=ラベル列, C=値列, A,D,E,F,G は外枠/装飾用
    widths = {"A": 2, "B": 22, "C": 60, "D": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 行高 (結合セルは自動拡張されないので手動設定)
    row_heights = {1: 18, 2: 14, 3: 50, 4: 14, 5: 60, 6: 14, 7: 14, 8: 14}
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    # タイトル (B3:C3)
    ws.merge_cells("B3:C3")
    title = ws["B3"]
    title.value = "EDC 仕様書"
    title.font = Font(name=BASE_FONT_NAME, size=28, bold=True, color="1F3864")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # 副題 / 試験名称 (B5:C5) — proper_name を中央大きめで
    ws.merge_cells("B5:C5")
    pn = ws["B5"]
    pn.value = study.proper_name
    pn.font = Font(name=BASE_FONT_NAME, size=12, color="404040")
    pn.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 区切り線 (B7:C7)
    ws.merge_cells("B7:C7")
    ws["B7"].border = Border(bottom=Side(style="medium", color="305496"))

    # メタ情報テーブル (B9〜)
    meta_rows = [
        ("試験 ID", study.name),
        ("発行日", date.today().isoformat()),
    ]
    start_row = 9
    for i, (k, v) in enumerate(meta_rows):
        r = start_row + i
        ws.row_dimensions[r].height = 22

        kc = ws.cell(row=r, column=2, value=k)
        kc.font = FONT_META_KEY
        kc.fill = FILL_META
        kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        kc.border = BORDER_ALL

        vc = ws.cell(row=r, column=3, value=v)
        vc.font = FONT_BASE
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        vc.border = BORDER_ALL

    # 承認欄 (B19:C21)
    sign_top = start_row + len(meta_rows) + 2
    ws.row_dimensions[sign_top].height = 18
    ws.merge_cells(start_row=sign_top, start_column=2, end_row=sign_top, end_column=3)
    sh = ws.cell(row=sign_top, column=2, value="承認")
    sh.font = FONT_BOLD
    sh.fill = FILL_META
    sh.alignment = Alignment(horizontal="center", vertical="center")
    sh.border = BORDER_ALL

    sign_headers = ["作成", "確認", "承認"]
    sign_row = sign_top + 1
    ws.row_dimensions[sign_row].height = 18
    ws.row_dimensions[sign_row + 1].height = 40
    # B列を作成、C列を 確認/承認 に分けるとレイアウトが崩れるため B/C を作成欄/確認欄として2分割
    layout = [
        (sign_row, 2, "作成"),
        (sign_row, 3, "確認・承認"),
    ]
    for r, c, label in layout:
        cell = ws.cell(row=r, column=c, value=label)
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    # 押印スペース
    for c in (2, 3):
        cell = ws.cell(row=sign_row + 1, column=c, value="")
        cell.border = BORDER_ALL

    # フッター (注釈)
    note_row = sign_row + 3
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=3)
    nc = ws.cell(row=note_row, column=2, value="本書は EDC 設定 JSON から EAGLE により自動生成されました。")
    nc.font = Font(name=BASE_FONT_NAME, size=8, italic=True, color="808080")
    nc.alignment = Alignment(horizontal="center", vertical="center")


# ---------------- 改訂履歴 ----------------
def _build_revision_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("改訂履歴", index=1)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "改訂履歴"
    ws["A1"].font = FONT_TITLE

    headers = ["版", "日付", "改訂者", "改訂内容"]
    widths = [8, 14, 16, 80]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = w

    # サンプル初版行
    sample = ["1.0", date.today().isoformat(), "", "初版作成 (EAGLE 自動生成)"]
    for i, v in enumerate(sample, start=1):
        c = ws.cell(row=4, column=i, value=v)
        c.font = FONT_BASE
        c.alignment = ALIGN_TOP_WRAP
        c.border = BORDER_ALL


# ---------------- 索引 ----------------
def _build_index_sheet(wb: Workbook, study: Study, sheet_names: list[str]) -> None:
    ws = wb.create_sheet("00_索引", index=2)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "索引"
    ws["A1"].font = FONT_TITLE

    meta_rows = [
        ("試験ID", study.name),
        ("試験名称", study.proper_name),
        ("疾患カテゴリ", study.disease_category or ""),
        ("実施機関", study.organization_name or ""),
        ("SDTM Version", study.sdtm_version or ""),
        ("SDTM Terminology", study.sdtm_terminology_version or ""),
        ("CTCAE Version", study.ctcae_version or ""),
        ("フォーム数", len(study.sheets)),
        ("総入力項目数", study.total_field_items()),
        ("コードリスト数 (使用中)", len(_used_codelist_names(study))),
        ("割付群数", len(study.sheet_groups)),
        ("必須チェック数", study.count_validator("presence")),
        ("日付範囲チェック数", study.count_validator("date")),
        ("数値範囲チェック数", study.count_validator("numericality")),
        ("ロジカルチェック数 (formula)", study.count_validator("formula")),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font = FONT_META_KEY
        kc.fill = FILL_META
        kc.alignment = ALIGN_LEFT
        kc.border = BORDER_ALL
        vc = ws.cell(row=i, column=2, value=v)
        vc.font = FONT_BASE
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        vc.border = BORDER_ALL

    # コードリスト一覧へのリンク
    cl_link_row = ws.max_row + 2
    link = ws.cell(row=cl_link_row, column=1, value="→ コードリスト一覧へ")
    link.hyperlink = "#'00_コードリスト一覧'!A1"
    link.font = FONT_LINK

    # フォーム一覧
    header_row = cl_link_row + 2
    for col, h in enumerate(INDEX_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL

    for i, (s, sname) in enumerate(zip(study.sheets, sheet_names), start=1):
        r = header_row + i
        for col, val in enumerate(
            [i, s.name, s.alias_name or "", s.category or "", len(s.field_items), "○" if s.is_closed else ""],
            start=1,
        ):
            c = ws.cell(row=r, column=col, value=val)
            c.font = FONT_BASE
            c.border = BORDER_ALL
            c.alignment = ALIGN_LEFT if col != 5 else ALIGN_CENTER
        link_cell = ws.cell(row=r, column=7, value=sname)
        link_cell.hyperlink = f"#'{sname}'!A1"
        link_cell.font = FONT_LINK
        link_cell.border = BORDER_ALL

    # 列幅
    for i, w in enumerate([24, 60, 20, 18, 8, 8, 36], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"


# ---------------- コードリスト一覧 ----------------
def _build_codelist_sheet(
    wb: Workbook, study: Study
) -> dict[str, int]:
    """使用中コードリストを書き出し、{codelist名: 開始行} を返す (フォーム側のリンク用)。"""
    ws = wb.create_sheet("00_コードリスト一覧", index=4)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "コードリスト一覧 (使用中のみ)"
    ws["A1"].font = FONT_TITLE

    back = ws["A2"]
    back.value = "← 索引へ戻る"
    back.hyperlink = "#'00_索引'!A1"
    back.font = FONT_LINK

    used = _used_codelist_names(study)
    # 使われている順 + 名前ソートで安定化
    codelists = sorted(
        [cl for cl in study.options if cl.name in used], key=lambda c: c.name
    )

    row = 3
    anchors: dict[str, int] = {}
    for cl in codelists:
        anchors[cl.name] = row
        # セクションヘッダー
        title = ws.cell(row=row, column=1, value=cl.name)
        title.font = Font(name=BASE_FONT_NAME, size=12, bold=True, color="FFFFFF")
        title.fill = FILL_CODELIST_HEADER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        # サブヘッダー (Usable 列削除)
        for col, h in enumerate(["Seq", "Code", "Name"], start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL
        row += 1

        # Usable=○ のみ書き出し
        for v in sorted([x for x in cl.values if x.is_usable], key=lambda x: x.seq):
            for col, val in enumerate([v.seq, v.code, v.name], start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONT_BASE
                c.border = BORDER_ALL
                # 全セル折り返しで隣セル溢れ防止
                if col == 1:
                    c.alignment = ALIGN_CENTER
                else:
                    c.alignment = ALIGN_TOP_WRAP
            row += 1
        row += 1  # 空行

    for i, w in enumerate([8, 32, 70], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    return anchors


# ---------------- 各フォームシート ----------------
def _build_form_sheet(
    wb: Workbook,
    sheet: Sheet,
    sheet_name: str,
    codelist_anchors: dict[str, int],
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    # A1: シート名 / A2: 索引へ戻る
    ws["A1"] = sheet.name
    ws["A1"].font = FONT_SHEET_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    back = ws["A2"]
    back.value = "← 索引へ戻る"
    back.hyperlink = "#'00_索引'!A1"
    back.font = FONT_LINK
    back.alignment = Alignment(horizontal="left", vertical="center")

    # A4 から表ヘッダー
    start_col = 1
    header_row = 4
    for i, h in enumerate(FORM_HEADERS):
        c = ws.cell(row=header_row, column=start_col + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(start_col + i)].width = FORM_COL_WIDTHS[i]

    sdtm_map = _build_sdtm_map(sheet)

    # データ (header_row + 1 行目から)
    for ridx, fi in enumerate(sheet.field_items, start=header_row + 1):
        num_min, num_max = _num_range(fi.validators)
        date_min, date_max = _date_range(fi.validators)
        fcond, fmsg = _formula(fi.validators)
        values = [
            fi.seq,
            fi.name,
            sdtm_map.get(fi.name, ""),
            fi.label,
            fi.type.replace("FieldItem::", ""),
            fi.field_type or "",
            _required(fi.validators),
            "" if fi.is_invisible else "○",
            fi.default_value or "",
            fi.option_name or "",
            num_min,
            num_max,
            date_min,
            date_max,
            fcond,
            fmsg,
            fi.reference_field or "",
            _description(fi),
        ]
        for i, val in enumerate(values):
            c = ws.cell(row=ridx, column=start_col + i, value=val)
            c.font = FONT_BASE
            c.border = BORDER_ALL
            # 全セル折り返し ON にして隣セルへの溢れを防止
            c.alignment = ALIGN_TOP_WRAP

        # Code List 列 (10 番目) をハイパーリンク化
        if fi.option_name and fi.option_name in codelist_anchors:
            cl_cell = ws.cell(row=ridx, column=start_col + 9)
            cl_cell.hyperlink = (
                f"#'00_コードリスト一覧'!A{codelist_anchors[fi.option_name]}"
            )
            cl_cell.font = FONT_LINK

    # ヘッダー行固定
    ws.freeze_panes = ws.cell(row=header_row + 1, column=start_col)


# ---------------- シートマトリクス (フォーム × 群) ----------------
def _build_matrix_sheet(wb: Workbook, study: Study, sheet_names: list[str]) -> None:
    """各フォームがどの sheet_group に属するかをマトリクス表示。"""
    ws = wb.create_sheet("00_シートマトリクス", index=3)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "シートマトリクス (フォーム × 割付群)"
    ws["A1"].font = FONT_TITLE
    back = ws["A2"]
    back.value = "← 索引へ戻る"
    back.hyperlink = "#'00_索引'!A1"
    back.font = FONT_LINK

    # 群ごとのメンバ alias 集合
    group_aliases: list[tuple[str, set[str]]] = [
        (g.name, {s.get("alias_name") for s in g.sheets}) for g in study.sheet_groups
    ]

    header_row = 4
    headers = ["No.", "Sheet Name", "Alias", "Items"] + [name for name, _ in group_aliases] + ["未分類"]
    widths = [6, 50, 20, 8] + [16] * len(group_aliases) + [10]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, (s, sname) in enumerate(zip(study.sheets, sheet_names), start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i).font = FONT_BASE
        link = ws.cell(row=r, column=2, value=s.name)
        link.hyperlink = f"#'{sname}'!A1"
        link.font = FONT_LINK
        ws.cell(row=r, column=3, value=s.alias_name or "").font = FONT_BASE
        ws.cell(row=r, column=4, value=len(s.field_items)).font = FONT_BASE

        any_group = False
        for j, (_, alias_set) in enumerate(group_aliases, start=5):
            mark = "○" if s.alias_name in alias_set else ""
            if mark:
                any_group = True
            c = ws.cell(row=r, column=j, value=mark)
            c.font = FONT_BASE
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL
        nc = ws.cell(row=r, column=5 + len(group_aliases), value="" if any_group else "○")
        nc.font = FONT_BASE
        nc.alignment = ALIGN_CENTER
        nc.border = BORDER_ALL

        for col in (1, 2, 3, 4):
            ws.cell(row=r, column=col).border = BORDER_ALL

    # 集計行
    summary_row = header_row + len(study.sheets) + 1
    ws.cell(row=summary_row, column=2, value="合計").font = FONT_BOLD
    for j, (_, alias_set) in enumerate(group_aliases, start=5):
        cnt = sum(1 for s in study.sheets if s.alias_name in alias_set)
        c = ws.cell(row=summary_row, column=j, value=cnt)
        c.font = FONT_BOLD
        c.alignment = ALIGN_CENTER
    unassigned = sum(
        1 for s in study.sheets if not any(s.alias_name in a for _, a in group_aliases)
    )
    c = ws.cell(row=summary_row, column=5 + len(group_aliases), value=unassigned)
    c.font = FONT_BOLD
    c.alignment = ALIGN_CENTER

    ws.freeze_panes = f"C{header_row + 1}"


# ---------------- エントリ ----------------
def build_spec_workbook(study: Study) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    used: set[str] = set()
    sheet_names = [_safe_sheet_name(s.name, used, i) for i, s in enumerate(study.sheets, start=1)]

    _build_cover_sheet(wb, study)
    _build_revision_sheet(wb)
    _build_index_sheet(wb, study, sheet_names)
    _build_matrix_sheet(wb, study, sheet_names)
    anchors = _build_codelist_sheet(wb, study)

    for s, sname in zip(study.sheets, sheet_names):
        _build_form_sheet(wb, s, sname, anchors)

    return wb


def write_spec_excel(study: Study, output_path: str | Path) -> Path:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_spec_workbook(study)
    wb.save(out)
    return out
