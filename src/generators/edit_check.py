"""エディットチェック確認書 (Excel) 生成。

4種の validator (presence / date / numericality / formula) を 1 シートに統合。
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..parser.models import FieldItem, Sheet, Study, Validators

BASE_FONT_NAME = "Meiryo UI"
FONT_BASE = Font(name=BASE_FONT_NAME, size=9)
FONT_BOLD = Font(name=BASE_FONT_NAME, size=10, bold=True)
FONT_HEADER = Font(name=BASE_FONT_NAME, size=10, bold=True, color="FFFFFF")
FONT_TITLE = Font(name=BASE_FONT_NAME, size=16, bold=True)

FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_REQ = PatternFill("solid", fgColor="E7E6E6")
FILL_DATE = PatternFill("solid", fgColor="DDEBF7")
FILL_NUM = PatternFill("solid", fgColor="FCE4D6")
FILL_FORMULA = PatternFill("solid", fgColor="E2EFDA")

THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_TOP_WRAP = Alignment(vertical="top", wrap_text=True)

CHECK_FILLS = {
    "Required": FILL_REQ,
    "Date": FILL_DATE,
    "Numericality": FILL_NUM,
    "Formula": FILL_FORMULA,
}

HEADERS = ["No.", "Sheet", "Field", "Label", "SDTM", "Check Type", "Condition", "Message"]
COL_WIDTHS = [6, 28, 12, 28, 14, 14, 50, 36]


def _sdtm_for(sheet: Sheet, field_name: str) -> str:
    out: list[str] = []
    for cfg in sheet.cdisc_sheet_configs:
        var = (cfg.table or {}).get(field_name)
        if var:
            out.append(f"{cfg.prefix}.{var}" if cfg.prefix else var)
    return ", ".join(out)


def _num_condition(v: Validators) -> str:
    n = v.numericality
    if not n:
        return ""
    parts = []
    if n.validate_numericality_greater_than_or_equal_to:
        parts.append(f">= {n.validate_numericality_greater_than_or_equal_to}")
    elif n.validate_numericality_greater_than:
        parts.append(f"> {n.validate_numericality_greater_than}")
    if n.validate_numericality_less_than_or_equal_to:
        parts.append(f"<= {n.validate_numericality_less_than_or_equal_to}")
    elif n.validate_numericality_less_than:
        parts.append(f"< {n.validate_numericality_less_than}")
    if n.validate_numericality_equal_to:
        parts.append(f"== {n.validate_numericality_equal_to}")
    return " かつ ".join(parts) if parts else "(範囲未設定)"


def _date_condition(v: Validators) -> str:
    d = v.date
    if not d:
        return ""
    parts = []
    if d.validate_date_after_or_equal_to:
        parts.append(f">= {d.validate_date_after_or_equal_to}")
    elif d.validate_date_after:
        parts.append(f"> {d.validate_date_after}")
    if d.validate_date_before_or_equal_to:
        parts.append(f"<= {d.validate_date_before_or_equal_to}")
    elif d.validate_date_before:
        parts.append(f"< {d.validate_date_before}")
    return " かつ ".join(parts) if parts else "(範囲未設定)"


def _iter_checks(study: Study):
    """(sheet, field, check_type, condition, message) を順次 yield。"""
    for sheet in study.sheets:
        for fi in sheet.field_items:
            v = fi.validators
            if v.presence is not None:
                yield sheet, fi, "Required", "値が入力されていること", "必須項目"
            if v.date is not None:
                cond = _date_condition(v)
                if cond:
                    yield sheet, fi, "Date", cond, "日付範囲外です"
            if v.numericality is not None:
                cond = _num_condition(v)
                if cond:
                    yield sheet, fi, "Numericality", cond, "数値範囲外です"
            if v.formula is not None and (v.formula.validate_formula_if or v.formula.validate_formula_message):
                yield (
                    sheet,
                    fi,
                    "Formula",
                    v.formula.validate_formula_if or "",
                    v.formula.validate_formula_message or "",
                )


# ---------------- 表紙 ----------------
def _build_cover(wb: Workbook, study: Study, total_checks: int) -> None:
    from openpyxl.worksheet.page import PageMargins

    ws = wb.create_sheet("表紙", index=0)
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.8, bottom=0.8)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    ws.row_dimensions[3].height = 50
    ws.merge_cells("B3:C3")
    t = ws["B3"]
    t.value = "エディットチェック確認書"
    t.font = Font(name=BASE_FONT_NAME, size=24, bold=True, color="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")

    # 試験名は長いことが多いので行高を確保 (結合セルは自動拡張されない)
    ws.row_dimensions[5].height = 60
    ws.merge_cells("B5:C5")
    pn = ws["B5"]
    pn.value = study.proper_name
    pn.font = Font(name=BASE_FONT_NAME, size=12, color="404040")
    pn.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("B7:C7")
    ws["B7"].border = Border(bottom=Side(style="medium", color="305496"))

    metas = [
        ("試験 ID", study.name),
        ("チェック総数", f"{total_checks:,}"),
        ("発行日", date.today().isoformat()),
    ]
    for i, (k, v) in enumerate(metas):
        r = 9 + i
        ws.row_dimensions[r].height = 22
        kc = ws.cell(row=r, column=2, value=k)
        kc.font = FONT_BOLD
        kc.fill = PatternFill("solid", fgColor="D9E1F2")
        kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        kc.border = BORDER_ALL
        vc = ws.cell(row=r, column=3, value=v)
        vc.font = FONT_BASE
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.border = BORDER_ALL


# ---------------- 改訂履歴 ----------------
def _build_revision(wb: Workbook) -> None:
    ws = wb.create_sheet("改訂履歴", index=1)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "改訂履歴"
    ws["A1"].font = FONT_TITLE
    headers = ["版", "日付", "改訂者", "改訂内容"]
    widths = [8, 14, 16, 80]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = w
    sample = ["1.0", date.today().isoformat(), "", "初版作成 (EAGLE 自動生成)"]
    for i, val in enumerate(sample, start=1):
        c = ws.cell(row=4, column=i, value=val)
        c.font = FONT_BASE
        c.alignment = ALIGN_TOP_WRAP
        c.border = BORDER_ALL


# ---------------- サマリ ----------------
def _build_summary(wb: Workbook, study: Study, counts: dict[str, int]) -> None:
    ws = wb.create_sheet("サマリ", index=2)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "チェック種別サマリ"
    ws["A1"].font = FONT_TITLE
    headers = ["Check Type", "件数"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    for i, key in enumerate(["Required", "Date", "Numericality", "Formula"], start=1):
        r = 3 + i
        kc = ws.cell(row=r, column=1, value=key)
        kc.font = FONT_BASE
        kc.fill = CHECK_FILLS[key]
        kc.border = BORDER_ALL
        kc.alignment = Alignment(horizontal="center")
        vc = ws.cell(row=r, column=2, value=counts.get(key, 0))
        vc.font = FONT_BASE
        vc.border = BORDER_ALL
        vc.alignment = Alignment(horizontal="right")
    # 合計
    total_r = 8
    tk = ws.cell(row=total_r, column=1, value="合計")
    tk.font = FONT_BOLD
    tk.border = BORDER_ALL
    tk.alignment = Alignment(horizontal="center")
    tv = ws.cell(row=total_r, column=2, value=sum(counts.values()))
    tv.font = FONT_BOLD
    tv.border = BORDER_ALL
    tv.alignment = Alignment(horizontal="right")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14


# ---------------- メインシート ----------------
def _build_check_sheet(wb: Workbook, study: Study) -> dict[str, int]:
    ws = wb.create_sheet("エディットチェック一覧")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "エディットチェック一覧"
    ws["A1"].font = FONT_TITLE

    header_row = 3
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]

    counts: dict[str, int] = {"Required": 0, "Date": 0, "Numericality": 0, "Formula": 0}
    no = 0
    for sheet, fi, ctype, cond, msg in _iter_checks(study):
        no += 1
        counts[ctype] += 1
        r = header_row + no
        values = [
            no,
            sheet.name,
            fi.name,
            fi.label,
            _sdtm_for(sheet, fi.name),
            ctype,
            cond,
            msg,
        ]
        for i, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = FONT_BASE
            c.border = BORDER_ALL
            c.alignment = ALIGN_TOP_WRAP if i in (4, 7, 8) else Alignment(
                vertical="top", wrap_text=True
            )
        # Check Type 列に色
        ws.cell(row=r, column=6).fill = CHECK_FILLS[ctype]

    # AutoFilter + 固定
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + no}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return counts


# ---------------- エントリ ----------------
def build_edit_check_workbook(study: Study) -> tuple[Workbook, dict[str, int]]:
    wb = Workbook()
    wb.remove(wb.active)
    counts = _build_check_sheet(wb, study)  # 末尾追加
    total = sum(counts.values())
    _build_cover(wb, study, total)
    _build_revision(wb)
    _build_summary(wb, study, counts)
    # シート順を再整列: 表紙, 改訂履歴, サマリ, エディットチェック一覧
    order = ["表紙", "改訂履歴", "サマリ", "エディットチェック一覧"]
    wb._sheets = [wb[n] for n in order]
    return wb, counts


def write_edit_check_excel(study: Study, output_path: str | Path) -> Path:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb, _ = build_edit_check_workbook(study)
    wb.save(out)
    return out
