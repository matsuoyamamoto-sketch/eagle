"""マニュアルチェックリスト (Excel) 生成 — Cohere 利用。"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..ai.cohere_client import CohereJSONClient
from ..ai.prompts import manual_check as P
from ..parser.models import Sheet, Study

BASE_FONT = "Meiryo UI"
F_BASE = Font(name=BASE_FONT, size=9)
F_HEADER = Font(name=BASE_FONT, size=10, bold=True, color="FFFFFF")
F_TITLE = Font(name=BASE_FONT, size=16, bold=True)
FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_HIGH = PatternFill("solid", fgColor="F8CBAD")
FILL_MID = PatternFill("solid", fgColor="FFE699")
FILL_LOW = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS = ["No.", "Sheet", "Category", "Target Fields", "Check Point", "Rationale", "Severity"]
WIDTHS = [6, 24, 16, 24, 50, 36, 10]


def _estimate_row_height(values: list, widths: list[int]) -> float:
    max_lines = 1
    for v, w in zip(values, widths):
        text = str(v) if v is not None else ""
        if not text:
            continue
        for line in text.replace("\r", "").split("\n"):
            est_chars = max(1, w - 1)
            wrap_lines = max(1, -(-int(len(line) * 1.5) // est_chars))
            max_lines = max(max_lines, wrap_lines)
    return max(15.0, min(150.0, max_lines * 14.0))


def _severity_fill(sev: str) -> PatternFill | None:
    return {"high": FILL_HIGH, "medium": FILL_MID, "low": FILL_LOW}.get(sev)


CATEGORY_ORDER = {"記入漏れ": 0, "単位・桁数": 1, "整合性": 2, "日付前後関係": 3, "重複入力": 4}
CROSS_SHEET_LABEL = "(横断)"


def generate_check_points(
    study: Study,
    selected_sheet_names: list[str],
    client: CohereJSONClient | None = None,
    on_progress: Callable[[int, int, str], None] | None = None,
) -> list[dict]:
    """マニュアルチェックリスト生成。

    AI はスタディ全体で 1 回だけ呼び出し『記入漏れ / 整合性』を抽出。
    『単位・桁数』および『横断』は決定論で生成する。
    """
    client = client or CohereJSONClient()
    out: list[dict] = []
    total_steps = 3
    if on_progress:
        on_progress(1, total_steps, "AIで記入漏れ・整合性を抽出中")

    # --- AI: 記入漏れ + 整合性 ---
    try:
        data = client.chat_json(
            P.SYSTEM,
            P.build_study_prompt(study, selected_sheet_names),
            P.SCHEMA,
        )
        sheet_field_map = {
            s.name: {fi.name for fi in s.field_items if fi.type != "FieldItem::Note"}
            for s in study.sheets
        }
        for cp in data.get("check_points", []):
            sheet_name = cp.get("sheet_name", "")
            if sheet_name not in sheet_field_map:
                continue
            tgt = cp.get("target_fields")
            if not tgt or (isinstance(tgt, list) and not [t for t in tgt if t]):
                continue
            tgt_str = " ".join(tgt) if isinstance(tgt, list) else str(tgt)
            referenced = re.findall(r"field\d+", tgt_str)
            if not referenced:
                continue
            if any(f not in sheet_field_map[sheet_name] for f in referenced):
                continue
            out.append(
                {
                    "sheet": sheet_name,
                    "category": cp.get("category", ""),
                    "target_fields": tgt,
                    "check_point": cp.get("check_point", ""),
                    "rationale": cp.get("rationale", ""),
                    "severity": cp.get("severity", "medium"),
                }
            )
    except Exception as e:
        out.append(
            {
                "sheet": "(error)",
                "category": "(error)",
                "target_fields": [],
                "check_point": f"AI生成エラー: {e}",
                "rationale": "",
                "severity": "low",
            }
        )

    # --- 決定論: 単位・桁数 ---
    if on_progress:
        on_progress(2, total_steps, "単位・桁数チェックを展開中")
    out.extend(generate_unit_digit_checks(study, selected_sheet_names))

    # --- 決定論: フォーム横断 ---
    if on_progress:
        on_progress(3, total_steps, "フォーム横断チェックを展開中")
    out.extend(generate_cross_sheet_checks(study, selected_sheet_names))

    # 並び順: シート順 → カテゴリ順 (横断は末尾)
    sheet_order = {s.name: i for i, s in enumerate(study.sheets)}
    out.sort(
        key=lambda cp: (
            len(sheet_order) if cp.get("sheet") == CROSS_SHEET_LABEL else sheet_order.get(cp.get("sheet", ""), 9999),
            CATEGORY_ORDER.get(cp.get("category", ""), 99),
        )
    )
    return out


def generate_unit_digit_checks(study: Study, selected_sheet_names: list[str]) -> list[dict]:
    """numericality 制約のある数値項目に対し、単位・桁数の目視確認チェックを生成。"""
    selected = set(selected_sheet_names)
    out: list[dict] = []
    for sheet in study.sheets:
        if sheet.name not in selected:
            continue
        for fi in sheet.field_items:
            if fi.type == "FieldItem::Note" or not fi.field_type:
                continue
            if fi.validators.numericality is None:
                continue
            out.append(
                {
                    "sheet": sheet.name,
                    "category": "単位・桁数",
                    "target_fields": [f"{fi.label}({fi.name})"],
                    "check_point": f"{fi.label} の単位・桁数の妥当性を確認する (例: mg/g 取り違え、桁ずれ)。",
                    "rationale": "numericality 範囲内でも単位・桁数の入力誤りはバリデーションで検出できないため。",
                    "severity": "medium",
                }
            )
    return out


def _sheet_signature(sheet: Sheet) -> tuple:
    """visit-replicated 判定用の構造シグネチャ (field_name + label の集合)。"""
    return tuple(sorted((fi.name, fi.label) for fi in sheet.field_items if fi.type != "FieldItem::Note"))


def generate_cross_sheet_checks(study: Study, selected_sheet_names: list[str]) -> list[dict]:
    """visit 順に並ぶ同構造シート (visit-replicated) の同項目について、
    日付の前後関係・重複入力 (数値/日付のみ) チェックを決定論的に生成する。"""
    selected = set(selected_sheet_names)
    # 同シグネチャ (= 親フォームのコピー) ごとに sheets をグルーピング
    groups: dict[tuple, list[Sheet]] = {}
    for sheet in study.sheets:
        if sheet.name not in selected:
            continue
        sig = _sheet_signature(sheet)
        if not sig:
            continue
        groups.setdefault(sig, []).append(sheet)

    checks: list[dict] = []
    for sheets in groups.values():
        if len(sheets) < 2:
            # 共通シート (症例登録票・同意取得など) は横断対象外
            continue
        # study.sheets の出現順 = visit 順 (parser がそのまま並べる前提)
        order_index = {s.name: study.sheets.index(s) for s in sheets}
        sheets_sorted = sorted(sheets, key=lambda s: order_index[s.name])
        # field_name + label が一致する項目を、最初のシートの field_items 順で走査
        first = sheets_sorted[0]
        for fi in first.field_items:
            if fi.type == "FieldItem::Note" or not fi.field_type:
                continue
            is_date = fi.field_type == "date"
            is_numeric = fi.validators.numericality is not None
            if not (is_date or is_numeric):
                continue
            targets = [f"{fi.label}({fi.name})[{s.name}]" for s in sheets_sorted]
            if is_date:
                checks.append(
                    {
                        "sheet": CROSS_SHEET_LABEL,
                        "category": "日付前後関係",
                        "target_fields": targets,
                        "check_point": "訪問順に日付が前後関係に矛盾なく入力されているか確認する。",
                        "rationale": "visit 間で同一項目の日付が逆転していないかを目視確認する。",
                        "severity": "high",
                    }
                )
            checks.append(
                {
                    "sheet": CROSS_SHEET_LABEL,
                    "category": "重複入力",
                    "target_fields": targets,
                    "check_point": "訪問間で同一の値が連続して入力されていないか (コピペ・記入誤りの疑い) を確認する。",
                    "rationale": "visit 間で値が完全一致する場合は転記ミスやコピー入力の可能性がある。",
                    "severity": "medium",
                }
            )
    return checks


def build_manual_check_workbook(study: Study, points: list[dict]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    cv = wb.create_sheet("表紙")
    cv.sheet_view.showGridLines = False
    cv.column_dimensions["A"].width = 2
    cv.column_dimensions["B"].width = 22
    cv.column_dimensions["C"].width = 60
    cv.merge_cells("B3:C3")
    t = cv["B3"]
    t.value = "マニュアルチェックリスト"
    t.font = Font(name=BASE_FONT, size=24, bold=True, color="1F3864")
    t.alignment = CENTER
    cv.row_dimensions[3].height = 50
    cv.row_dimensions[5].height = 60
    cv.merge_cells("B5:C5")
    pn = cv["B5"]
    pn.value = study.proper_name
    pn.font = Font(name=BASE_FONT, size=12, color="404040")
    pn.alignment = CENTER
    metas = [("試験 ID", study.name), ("チェック件数", f"{len(points):,}"), ("発行日", date.today().isoformat())]
    for i, (k, v) in enumerate(metas):
        r = 9 + i
        cv.row_dimensions[r].height = 22
        kc = cv.cell(row=r, column=2, value=k)
        kc.font = Font(name=BASE_FONT, size=10, bold=True)
        kc.fill = PatternFill("solid", fgColor="D9E1F2")
        kc.border = BORDER_ALL
        kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc = cv.cell(row=r, column=3, value=v)
        vc.font = F_BASE
        vc.border = BORDER_ALL
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws = wb.create_sheet("マニュアルチェック一覧")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "マニュアルチェック一覧"
    ws["A1"].font = F_TITLE

    header_row = 3
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]

    for idx, cp in enumerate(points, start=1):
        r = header_row + idx
        targets = cp.get("target_fields") or []
        if isinstance(targets, list):
            targets = ", ".join(targets)
        values = [
            idx,
            cp.get("sheet", ""),
            cp.get("category", ""),
            targets,
            cp.get("check_point", ""),
            cp.get("rationale", ""),
            cp.get("severity", ""),
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = F_BASE
            c.border = BORDER_ALL
            c.alignment = WRAP
        fill = _severity_fill(cp.get("severity", ""))
        if fill is not None:
            ws.cell(row=r, column=7).fill = fill
        ws.row_dimensions[r].height = _estimate_row_height(values, WIDTHS)

    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(points)}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return wb


def write_manual_check_excel(
    study: Study,
    selected_sheet_names: list[str],
    output_path: str | Path,
    client: CohereJSONClient | None = None,
    on_progress: Callable[[int, int, str], None] | None = None,
) -> Path:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    points = generate_check_points(study, selected_sheet_names, client, on_progress)
    wb = build_manual_check_workbook(study, points)
    wb.save(out)
    return out
